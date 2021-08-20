import pandas as pd
from openpyxl import load_workbook
import time
import xlwt
# file = 'table.xlsx'

arr_EN = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

# Assign spreadsheet filename to `file`


class AmwayParser():
    def get_file(self):
        print("Введите имя файла с расширением. Пример: (table.xlsx)")
        print("Внимание! Файл должен находиться в директории с программой!")
        file = input("$:")
        try:
            wb = load_workbook(file)
            self.wb = wb
            return wb
        except Exception as ex:
            print("Вы ввели неверное имя файла! Программа завершила свою работу!")
            time.sleep(3)
            raise SystemExit(1)

    def get_sheet(self):
            print("Введите название листа, который желаете спарсить.")
            print(f"Список доступных листов:", end=" ")

            for i in self.wb.sheetnames:
                print(f"{i}", end=" ")
            print()

            sheetName = input("$:")
            try:
                sheet = self.wb[sheetName]
                self.sheet = sheet
                return sheet

            except Exception as ex:
                print("Неверное название листа. Программа завершилась с ошибкой.")
                time.sleep(3)
                raise SystemExit(1)

    def get_letter(self):
        print("Введите букву колонки, по которой должен проходить поиск:")
        letter = input("$:")
        self.letter = letter
        return letter

    def get_word(self):
        print("Введите ключевое слово, по которому хотите искать пользователей:")
        word = input("$:")
        self.word = word
        return word

    def search(self):
        max_rows = self.sheet.max_row
        
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("Python Sheet 1") 
        writed_rows = 0
        for i in range(max_rows):
            if self.sheet[f"{self.letter}{i+1}"].value == self.word:
                for j in range(self.sheet.max_column):
                    sheet1.write(writed_rows, j, self.sheet[f"{arr_EN[j]}{i+1}"].value) 
                    # print(self.sheet[f"{arr_EN[j]}{i+1}"].value, end=" ")
                writed_rows += 1
                print()
        book.save("result.xlsx")
        return True